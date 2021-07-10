from atsim.potentials.config import Configuration
from atsim.potentials.pair_tabulation import Excel_PairTabulation
from atsim.potentials.eam_tabulation import Excel_EAMTabulation
from atsim.potentials.eam_tabulation import Excel_FinnisSinclair_EAMTabulation
from atsim.potentials.eam_tabulation import Excel_ADP_EAMTabulation

from openpyxl import load_workbook, Workbook

import pytest

import io


def test_pair_model(tmp_path):

    cfg = u"""[Tabulation]
target : excel
dr : 0.01
cutoff : 5

[Pair]
O-O : as.polynomial 0 1
Al-O : as.polynomial 0 2

"""

    infile = io.StringIO(cfg)
    configuration = Configuration()
    tabulation = configuration.read(infile)

    assert type(tabulation) is Excel_PairTabulation

    out_path = tmp_path / "pair.xlsx"

    with tabulation.open_fp(out_path) as outfile:
        tabulation.write(outfile)

    wb = load_workbook(out_path)

    assert ["Pair"] == wb.sheetnames
    ws = wb["Pair"]

    # Check column headings
    assert ["r", "Al-O", "O-O"] == [c.value for c in next(ws.rows)]

    # Check number of table rows
    assert len(next(ws.columns)) == 502

    # Check some values
    assert ws["A2"].value == 0.0
    assert ws["B2"].value == 0.0
    assert ws["C2"].value == 0.0

    assert ws["A3"].value == 0.01
    assert ws["B3"].value == pytest.approx(2 * 0.01)
    assert ws["C3"].value == pytest.approx(0.01)

    assert ws["A502"].value == 5.0
    assert ws["B502"].value == pytest.approx(2.0 * 5.0)
    assert ws["C502"].value == pytest.approx(5.0)


def test_pair_eam(tmp_path):

    cfg = u"""[Tabulation]
target : excel_eam
dr : 0.01
cutoff : 5
drho : 0.01
cutoff_rho : 5

[Pair]
O-O : as.polynomial 0 1
Al-O : as.polynomial 0 2

[EAM-Density]
O : as.polynomial 0 3
Al : as.polynomial 0 4

[EAM-Embed]
O : as.polynomial 0 5
Al : as.polynomial 0 6

"""

    infile = io.StringIO(cfg)
    configuration = Configuration()
    tabulation = configuration.read(infile)

    assert type(tabulation) is Excel_EAMTabulation

    out_path = tmp_path / "pair.xlsx"

    with tabulation.open_fp(out_path) as outfile:
        tabulation.write(outfile)

    wb = load_workbook(out_path)

    assert ["Pair", "EAM-Density", "EAM-Embed"] == wb.sheetnames

    # Check [Pair] table
    ws = wb["Pair"]

    # Check column headings
    assert ["r", "Al-O", "O-O"] == [c.value for c in next(ws.rows)]

    # Check number of table rows
    assert len(next(ws.columns)) == 502

    # Check some values
    assert ws["A2"].value == 0.0
    assert ws["B2"].value == 0.0
    assert ws["C2"].value == 0.0

    assert ws["A3"].value == 0.01
    assert ws["B3"].value == pytest.approx(2 * 0.01)
    assert ws["C3"].value == pytest.approx(0.01)

    assert ws["A502"].value == 5.0
    assert ws["B502"].value == pytest.approx(2.0 * 5.0)
    assert ws["C502"].value == pytest.approx(5.0)

    # Check [EAM-Density] table
    ws = wb["EAM-Density"]

    # Check column headings
    assert ["r", "Al", "O"] == [c.value for c in next(ws.rows)]

    # Check number of table rows
    assert len(next(ws.columns)) == 502

    # Check some values
    assert ws["A2"].value == 0.0
    assert ws["B2"].value == 0.0
    assert ws["C2"].value == 0.0

    assert ws["A3"].value == 0.01
    assert ws["B3"].value == pytest.approx(4 * 0.01)
    assert ws["C3"].value == pytest.approx(3 * 0.01)

    assert ws["A502"].value == 5.0
    assert ws["B502"].value == pytest.approx(4.0 * 5.0)
    assert ws["C502"].value == pytest.approx(3.0 * 5.0)

    # Check [EAM-Embed] table
    ws = wb["EAM-Embed"]

    # Check column headings
    assert ["rho", "Al", "O"] == [c.value for c in next(ws.rows)]

    # Check number of table rows
    assert len(next(ws.columns)) == 502

    # Check some values
    assert ws["A2"].value == 0.0
    assert ws["B2"].value == 0.0
    assert ws["C2"].value == 0.0

    assert ws["A3"].value == 0.01
    assert ws["B3"].value == pytest.approx(6 * 0.01)
    assert ws["C3"].value == pytest.approx(5 * 0.01)

    assert ws["A502"].value == 5.0
    assert ws["B502"].value == pytest.approx(6.0 * 5.0)
    assert ws["C502"].value == pytest.approx(5.0 * 5.0)


def test_pair_eam_fs(tmp_path):

    cfg = u"""[Tabulation]
target : excel_eam_fs
dr : 0.01
cutoff : 5
drho : 0.01
cutoff_rho : 5

[Pair]
O-O : as.polynomial 0 1
Al-O : as.polynomial 0 2

[EAM-Density]
Al->O : as.polynomial 0 3
Al->Al : as.polynomial 0 4

[EAM-Embed]
O : as.polynomial 0 5
Al : as.polynomial 0 6

"""

    infile = io.StringIO(cfg)
    configuration = Configuration()
    tabulation = configuration.read(infile)

    assert type(tabulation) is Excel_FinnisSinclair_EAMTabulation

    out_path = tmp_path / "pair.xlsx"

    with tabulation.open_fp(out_path) as outfile:
        tabulation.write(outfile)

    wb = load_workbook(out_path)

    assert ["Pair", "EAM-Density", "EAM-Embed"] == wb.sheetnames

    # Check [Pair] table
    ws = wb["Pair"]

    # Check column headings
    assert ["r", "Al-O", "O-O"] == [c.value for c in next(ws.rows)]

    # Check number of table rows
    assert len(next(ws.columns)) == 502

    # Check some values
    assert ws["A2"].value == 0.0
    assert ws["B2"].value == 0.0
    assert ws["C2"].value == 0.0

    assert ws["A3"].value == 0.01
    assert ws["B3"].value == pytest.approx(2 * 0.01)
    assert ws["C3"].value == pytest.approx(0.01)

    assert ws["A502"].value == 5.0
    assert ws["B502"].value == pytest.approx(2.0 * 5.0)
    assert ws["C502"].value == pytest.approx(5.0)

    # Check [EAM-Density] table
    ws = wb["EAM-Density"]

    # Check column headings
    assert ["r", "Al->Al", "Al->O", "O->Al", "O->O"] == [c.value for c in next(ws.rows)]

    # Check number of table rows
    assert len(next(ws.columns)) == 502

    # Check some values
    assert ws["A2"].value == 0.0
    assert ws["B2"].value == 0.0
    assert ws["C2"].value == 0.0
    assert ws["D2"].value == 0.0
    assert ws["E2"].value == 0.0

    assert ws["A3"].value == 0.01
    assert ws["B3"].value == pytest.approx(4.0 * 0.01)
    assert ws["C3"].value == pytest.approx(3.0 * 0.01)
    assert ws["D3"].value == 0.0
    assert ws["E3"].value == 0.0

    # Check [EAM-Embed] table
    ws = wb["EAM-Embed"]

    # Check column headings
    assert ["rho", "Al", "O"] == [c.value for c in next(ws.rows)]

    # Check number of table rows
    assert len(next(ws.columns)) == 502

    # Check some values
    assert ws["A2"].value == 0.0
    assert ws["B2"].value == 0.0
    assert ws["C2"].value == 0.0

    assert ws["A3"].value == 0.01
    assert ws["B3"].value == pytest.approx(6 * 0.01)
    assert ws["C3"].value == pytest.approx(5 * 0.01)

    assert ws["A502"].value == 5.0
    assert ws["B502"].value == pytest.approx(6.0 * 5.0)
    assert ws["C502"].value == pytest.approx(5.0 * 5.0)


def test_excel_eam_adp(tmp_path):

    cfg = u"""[Tabulation]
target : excel_eam_adp
dr : 0.01
cutoff : 5
drho : 0.01
cutoff_rho : 5

[Pair]
O-O : as.polynomial 0 1
Al-O : as.polynomial 0 2

[EAM-Density]
Al : as.polynomial 0 3
O : as.polynomial 0 4

[EAM-Embed]
O : as.polynomial 0 5
Al : as.polynomial 0 6

[EAM-ADP-Dipole]
Al-O : as.polynomial 0 7
O-O : as.polynomial 0 8

[EAM-ADP-Quadrupole]
Al-Al : as.polynomial 0 9

"""

    infile = io.StringIO(cfg)
    configuration = Configuration()
    tabulation = configuration.read(infile)

    assert type(tabulation) is Excel_ADP_EAMTabulation

    out_path = tmp_path / "pair.xlsx"

    with tabulation.open_fp(out_path) as outfile:
        tabulation.write(outfile)

    wb = load_workbook(out_path)

    assert ["Pair", "EAM-Density", "EAM-Embed", "EAM-ADP-Dipole", "EAM-ADP-Quadrupole"] == wb.sheetnames

    # Check [Pair] table
    ws = wb["Pair"]

    # Check column headings
    assert ["r", "Al-O", "O-O"] == [c.value for c in next(ws.rows)]

    # Check number of table rows
    assert len(next(ws.columns)) == 502

    # Check some values
    assert ws["A2"].value == 0.0
    assert ws["B2"].value == 0.0
    assert ws["C2"].value == 0.0

    assert ws["A3"].value == 0.01
    assert ws["B3"].value == pytest.approx(2 * 0.01)
    assert ws["C3"].value == pytest.approx(0.01)

    assert ws["A502"].value == 5.0
    assert ws["B502"].value == pytest.approx(2.0 * 5.0)
    assert ws["C502"].value == pytest.approx(5.0)

    # Check [EAM-Density] table
    ws = wb["EAM-Density"]

    # Check column headings
    assert ["r", "Al", "O"] == [c.value for c in next(ws.rows)]

    # Check number of table rows
    assert len(next(ws.columns)) == 502

    # Check some values
    assert ws["A2"].value == 0.0
    assert ws["B2"].value == 0.0
    assert ws["C2"].value == 0.0

    assert ws["A3"].value == 0.01
    assert ws["B3"].value == pytest.approx(3.0 * 0.01)
    assert ws["C3"].value == pytest.approx(4.0 * 0.01)

    # Check [EAM-Embed] table
    ws = wb["EAM-Embed"]

    # Check column headings
    assert ["rho", "Al", "O"] == [c.value for c in next(ws.rows)]

    # Check number of table rows
    assert len(next(ws.columns)) == 502

    # Check some values
    assert ws["A2"].value == 0.0
    assert ws["B2"].value == 0.0
    assert ws["C2"].value == 0.0

    assert ws["A3"].value == 0.01
    assert ws["B3"].value == pytest.approx(6 * 0.01)
    assert ws["C3"].value == pytest.approx(5 * 0.01)

    assert ws["A502"].value == 5.0
    assert ws["B502"].value == pytest.approx(6.0 * 5.0)
    assert ws["C502"].value == pytest.approx(5.0 * 5.0)

    # Check [EAM-Dipole] table
    ws = wb["EAM-ADP-Dipole"]

    # Check column headings
    assert ["r", "Al-O", "O-O"] == [c.value for c in next(ws.rows)]

    # Check number of table rows
    assert len(next(ws.columns)) == 502

    # Check some values
    assert ws["A2"].value == 0.0
    assert ws["B2"].value == 0.0
    assert ws["C2"].value == 0.0

    assert ws["A3"].value == 0.01
    assert ws["B3"].value == pytest.approx(7 * 0.01)
    assert ws["C3"].value == pytest.approx(8 * 0.01)

    assert ws["A502"].value == 5.0
    assert ws["B502"].value == pytest.approx(7.0 * 5.0)
    assert ws["C502"].value == pytest.approx(8.0 * 5.0)
    
    # Check [EAM-Quadrupole] table
    ws = wb["EAM-ADP-Quadrupole"]

    # Check column headings
    assert ["r", "Al-Al"] == [c.value for c in next(ws.rows)]

    # Check number of table rows
    assert len(next(ws.columns)) == 502

    # Check some values
    assert ws["A2"].value == 0.0
    assert ws["B2"].value == 0.0

    assert ws["A3"].value == 0.01
    assert ws["B3"].value == pytest.approx(9 * 0.01)

    assert ws["A502"].value == 5.0
    assert ws["B502"].value == pytest.approx(9 * 5.0)

