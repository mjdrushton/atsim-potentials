from ._lammps_writeTABLE import writePotentials as lmp_writePotentials
from ._dlpoly_writeTABLE import writePotentials as dlpoly_writePotentials

def _r_value_iterator(tabulation):
  #for n in range(tabulation.nr+1):
  for n in range(tabulation.nr):
    yield float(n)* tabulation.cutoff / (float(tabulation.nr) -1)



class PairTabulation_AbstractBase(object):
  """Base class for PairTabulation objects. 

  Child classes must implement: 
    write() method
    """

  def __init__(self, potentials, cutoff, nr, target):
    """Create pair tabulation for LAMMPS.

    :param potentials: List of atsim.potentials.Potential objects.
    :param cutoff: Maximum separation to be tabulated.
    :param nr: Number of points to be used in tabulation.
    :param target: Label identifying the code for which this object creates tables."""
    self._nr = nr
    self._cutoff = cutoff
    self._potentials = potentials
    self._target = target
  
  @property
  def type(self):
    return "Pair"

  @property
  def target(self):
    return self._target

  @property
  def nr(self):
    return self._nr

  @property
  def cutoff(self):
    return self._cutoff

  @property
  def potentials(self):
    return self._potentials

  @property
  def dr(self):
    return (self.cutoff / float(self.nr-1)) 

  @classmethod
  def open_fp(self, filename):
    """Creates a file object with a given path suitable for writing potential data to.

    :param filename: Filename of output file object.

    :return: File object suitable for passing to write() method"""
    return open(filename, 'w')

  def write(self, fp):
    """Write the tabulation to the file object `fp`.

    :param fp: File object into which data should be written."""
    raise NotImplementedError("Sub-classes must implement write method")


class LAMMPS_PairTabulation(PairTabulation_AbstractBase):
  """Class for tabulating pair-potential models for LAMMPS"""
  
  def __init__(self, potentials, cutoff, nr):
    """Create pair tabulation for LAMMPS.

    :params potentials: List of atsim.potentials.Potential objects.
    :params cutoff: Maximum separation to be tabulated.
    :params nr: Number of points to be used in tabulation"""
    super(LAMMPS_PairTabulation, self).__init__(potentials, cutoff, nr, u"LAMMPS")

  def write(self, fp):
    """Write the tabulation to the file object `fp`.

    :param fp: File object into which data should be written."""
    lmp_writePotentials(self.potentials, self.dr, self.cutoff, self.nr-1, fp)


class DLPoly_PairTabulation(PairTabulation_AbstractBase):
  """Class for tabulating pair-potential models for DLPOLY"""

  def __init__(self, potentials, cutoff, nr):
    """Create pair tabulation for DLPOLY.

    :params potentials: List of atsim.potentials.Potential objects.
    :params cutoff: Maximum separation to be tabulated.
    :params nr: Number of points to be used in tabulation"""
    super(DLPoly_PairTabulation, self).__init__(potentials, cutoff, nr, u"DLPOLY")

  def write(self, fp):
    """Write tabulation to the file object `fp`.

    :param fp: File object into which data should be written."""
    dlpoly_writePotentials(self.potentials, self.cutoff, self.nr, fp)


class GULP_PairTabulation(PairTabulation_AbstractBase):
  """Class for tabulating pair-potential models for the GULP code.

  .. :seealso::

      * `Gulp Web site <https://nanochemistry.curtin.edu.au/gulp/>`_

  """

  def __init__(self, potentials, cutoff, nr):
    """Create pair tabulation for GULP.

    :params potentials: List of atsim.potentials.Potential objects.
    :params cutoff: Maximum separation to be tabulated.
    :params nr: Number of points to be used in tabulation"""
    super(GULP_PairTabulation, self).__init__(potentials, cutoff, nr, u"GULP")

  def write(self, fp):
    """Write tabulation to the file object `fp`.

    :param fp: File object into which data should be written."""
    
    for pot in self.potentials:
      self._write_pot(pot, fp)

  def _write_pot(self, pot, fp):
    header_template = u"{speciesA} {speciesB} {cutoff}\n"
    row_template = u"{energy:.10f} {sepn:.10f}\n"

    fp.write(u"spline cubic\n")
    fp.write(header_template.format(speciesA = pot.speciesA, speciesB = pot.speciesB, cutoff= self.cutoff))

    for r in _r_value_iterator(self):
      energy = pot.energy(r)
      fp.write(row_template.format(sepn = r, energy = energy))
    # fp.write(u"\n")


class Excel_PairTabulation(PairTabulation_AbstractBase):
  """Class for dumping pair-potential models into an Excel formatted spreadsheet"""

  def __init__(self, potentials, cutoff, nr):
    """Create pair tabulation for Excel.

    :params potentials: List of atsim.potentials.Potential objects.
    :params cutoff: Maximum separation to be tabulated.
    :params nr: Number of points to be used in tabulation"""
    super(Excel_PairTabulation, self).__init__(potentials, cutoff, nr, u"excel")
    self._workbook = None

  @property
  def workbook(self):
    """Property which returns an openpyxl.Workbook instance containing potential data"""
    if self._workbook is None:
      self._workbook = self._build_workbook()
    return self._workbook

  def _build_workbook(self):
    from openpyxl import Workbook
    wb = Workbook()
    wb.remove(wb.active)
    self._add_worksheets(wb)
    return wb

  def _populate_worksheet(self, ws, first_col_name, first_col_values, column_keys, column_dict):
    ws["A1"] = first_col_name
    for label, col in zip(column_keys, ws.iter_cols(min_row=1, max_row=1, min_col=2, max_col=len(column_keys)+1)):
      col[0].value = label

    for r_idx, r in enumerate(first_col_values):
      r_idx += 2
      ws.cell(r_idx, 1, value=r)
      for label, col in zip(column_keys, ws.iter_cols(min_row=r_idx, max_row=r_idx, min_col=2, max_col=len(column_keys)+1)):
        pot = column_dict[label]
        col[0].value = pot(r)


  def _add_pair_worksheet(self, wb):
    ws = wb.create_sheet("Pair")

    # Get sorted list of potentials
    pot_dict = {}
    for p in self.potentials:
      k = "{}-{}".format(*sorted([p.speciesA, p.speciesB]))
      v = p.potentialFunction
      pot_dict[k] = v
    column_heads = sorted(pot_dict.keys())
    self._populate_worksheet(ws, "r", _r_value_iterator(self), column_heads, pot_dict )


  def _add_worksheets(self, wb):
    self._add_pair_worksheet(wb)

  def write(self, fp):
    """Write tabulation to the file object `fp` (note: fp should be opened in binary mode).

    :param fp: File object into which data should be written."""
    wb = self.workbook

    from tempfile import NamedTemporaryFile
    with NamedTemporaryFile() as tmp:
        wb.save(tmp.name)
        tmp.seek(0)
        fp.write(tmp.read())

  @classmethod
  def open_fp(self, filename):
    """Creates a file object with a given path suitable for writing potential data to.

    :param filename: Filename of output file object.

    :return: File object suitable for passing to write() method"""
    return open(filename, 'wb')

